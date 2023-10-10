import geocoder
import requests
from bs4 import BeautifulSoup
from geopy.geocoders import Nominatim
from PIL import Image
import wget
import zipfile
import os
im = Image.open("map2.png").convert('RGB')
import math
import openpyxl
from pyexcel.cookbook import merge_all_to_a_book
import glob

def return_longlat():
    g = geocoder.ip('me')
    latitude, longitude = g.latlng
    return latitude, longitude

def longlat_to_city(lat, lon):
    geolocator = Nominatim(user_agent="geoapiExercises")
    location = geolocator.reverse(str(lat)+","+str(lon))
    print(location)

def weather(city, province): 
    # creating url and requests instance
    url = "https://www.google.com/search?q="+"weather"+city+province
    html = requests.get(url).content
    
    # getting raw data
    soup = BeautifulSoup(html, 'html.parser')
    temp = soup.find('div', attrs={'class': 'BNeawe iBp4i AP7Wnd'}).text

    # printing all data
    return temp


def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371  # Earth's radius in kilometers
    
    # Convert coordinates to radians
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    
    # Calculate differences in coordinates
    delta_lat = lat2_rad - lat1_rad
    delta_lon = lon2_rad - lon1_rad
    
    # Apply Haversine formula
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    distance = R * c
    
    return distance

def find_nearest_location(reference_point, locations):
    min_distance = float('inf')
    nearest_location = None
    
    for location in locations:
        if (location['latitude'] != None and location['longitude'] != None):
            distance = haversine_distance(reference_point['latitude'], reference_point['longitude'], location['latitude'], location['longitude'])
        
        if distance < min_distance:
            min_distance = distance
            nearest_location = location
    
    return nearest_location

# # Example usage
# reference_point = {'latitude': 37.7749, 'longitude': -122.4194}
# locations = [
#     {'latitude': 37.7749, 'longitude': -122.4194},
#     {'latitude': 34.0522, 'longitude': -118.2437},
#     {'latitude': 40.7128, 'longitude': -74.0060},
# ]

# nearest_location = find_nearest_location(reference_point, locations)
# print(nearest_location)


def return_longlat():
    g = geocoder.ip('me')
    latitude, longitude = g.latlng
    # print(f"Latitude: {latitude}, Longitude: {longitude}")
    return latitude, longitude


def findClosestLake(latitude, longitude): 
    referencePoint = {'latitude': latitude, 'longitude': longitude}
    locations = []
    lakeCoords = []
    wb = openpyxl.load_workbook('station.xlsx')
    ws = wb.active
    lofname = []
    loflongitude = []
    loflatitude = []
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=4).value
        lofname.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=13).value
        loflongitude.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=12).value
        loflatitude.append(data)
    
    for i in range(0, len(loflongitude)):
        x = {'latitude': loflatitude[i], 'longitude': loflongitude[i]}
        locations.append(x)
    
    for i in range(0, len(lofname)):
        key = str(loflatitude[i]) + ',' + str(loflongitude[i])
        x = {key: lofname[i]}
        lakeCoords.append(x)
        
    nearest_location = find_nearest_location(referencePoint, locations)
    # print(nearest_location)
    x = nearest_location['latitude']
    y = nearest_location['longitude']
    key = str(x) + ',' + str(y)
    for i in range(0, len(lakeCoords)):
        if key in lakeCoords[i]:
            name = (lakeCoords[i])[key]
            break
    return name


def findClosestCity(latitude, longitude): 
    referencePoint = {'latitude': latitude, 'longitude': longitude}
    locations = []
    cityCoords = []
    wb = openpyxl.load_workbook('canadacities.xlsx')
    ws = wb.active
    lofname = []
    loftype = []
    loflongitude = []
    loflatitude = []
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=2).value
        lofname.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=3).value
        loftype.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=6).value
        loflongitude.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=5).value
        loflatitude.append(data)
    
    for i in range(0, len(loflongitude)):
        x = {'latitude': loflatitude[i], 'longitude': loflongitude[i]}
        locations.append(x)
    
    for i in range(0, len(lofname)):
        key = str(loflatitude[i]) + ',' + str(loflongitude[i])
        x = {key: lofname[i]}
        cityCoords.append(x)
        
    nearest_location = find_nearest_location(referencePoint, locations)
    # print(nearest_location)
    x = nearest_location['latitude']
    y = nearest_location['longitude']
    key = str(x) + ',' + str(y)
    for i in range(0, len(cityCoords)):
        if key in cityCoords[i]:
            name = (cityCoords[i])[key]
            prov = (cityCoords[i])[key]
            break
    for j in range(2, ws.max_row+1):
        data = ws.cell(j,column=2).value
        if(data==name):
            prov = ws.cell(j,column=3).value
    return name, prov



def drainageToLake(latitude, longitude): 
    name = findClosestLake(latitude, longitude)
    wb = openpyxl.load_workbook('station.xlsx')
    ws = wb.active
    lofname = []
    lofdrainage= []
    list = []
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=4).value
        lofname.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=8).value
        lofdrainage.append(data)
    for i in range(0, len(lofdrainage)):
        x = {str(lofname[i]): lofdrainage[i]}
        list.append(x)
    for i in range(0, len(list)):
        if name in list[i]:
            if (list[i])[name] != None:
                drainage = str((list[i])[name]) + " sq mi"
                break
            else:
                drainage = "Unknown Amount"
    return drainage, name


def typeOfWater(latitude, longitude): 
    name = findClosestLake(latitude, longitude)
    wb = openpyxl.load_workbook('station.xlsx')
    ws = wb.active
    lofname = []
    loftype= []
    list = []
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=4).value
        lofname.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=5).value
        loftype.append(data)
    for i in range(0, len(loftype)):
        x = {str(lofname[i]): loftype[i]}
        list.append(x)
    for i in range(0, len(list)):
        if name in list[i]:
            if (list[i])[name] != None:
                type = str((list[i])[name])
                break
            else:
                type = "Unknown Water Source"
    return type

def redownload_lakes_excel():
    wget.download("https://www.waterqualitydata.us/data/Station/search?countrycode=CA&mimeType=csv&zip=yes&providers=NWIS&providers=STEWARDS&providers=STORET")
    with zipfile.ZipFile("station.zip", "r") as zip_ref:
        zip_ref.extractall()
    os.remove("station.zip")
    merge_all_to_a_book(glob.glob("station.csv"), "station.xlsx")
    os.remove("station.csv")

def return_ecosystem(return_longlat):
    if im.getpixel((return_longlat()[0] + 100, return_longlat()[1] + 140)) == (74, 161, 112):
        eco_system = "Evergreen ecosystem"
    if im.getpixel((return_longlat()[0] + 100, return_longlat()[1] + 140)) == (240, 225, 117):
        eco_system = "Deciduous ecosystem"
    if im.getpixel((return_longlat()[0] + 100, return_longlat()[1] + 140)) == (228, 237, 251):
        eco_system = "Shrubland"
    if im.getpixel((return_longlat()[0] + 100, return_longlat()[1] + 140)) == (192, 203, 149):
        eco_system = "grassland"
    if im.getpixel((return_longlat()[0] + 100, return_longlat()[1] + 140)) == (205, 213, 230):
        eco_system = "Cropland"
    if im.getpixel((return_longlat()[0] + 100, return_longlat()[1] + 140)) == (252, 252, 252):
        eco_system = "Cropland"
    if im.getpixel((return_longlat()[0] + 100, return_longlat()[1] + 140)) == (74, 161, 112):
        eco_system = "Built-in"
    print(eco_system)
    return eco_system


def find_closest_animals(latitude, longitude):
    referencePoint = {'latitude': latitude, 'longitude': longitude}
    locations = []
    cityCoords = []

    wb = openpyxl.load_workbook('animals.xlsx')
    ws = wb.active
    lofname = []
    loflongitude = []
    loflatitude = []
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=1).value
        lofname.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=2).value
        loflongitude.append(data)
    for i in range(2, ws.max_row+1):
        data = ws.cell(i,column=3).value
        loflatitude.append(data)
    
    for i in range(0, len(loflongitude)):
        x = {'latitude': loflatitude[i], 'longitude': loflongitude[i]}
        locations.append(x)
    
    for i in range(0, len(lofname)):
        key = str(loflatitude[i]) + ',' + str(loflongitude[i])
        x = {key: lofname[i]}
        cityCoords.append(x)
        
    nearest_location = find_nearest_location(referencePoint, locations)
    # print(nearest_location)
    x = nearest_location['latitude']
    y = nearest_location['longitude']
    key = str(x) + ',' + str(y)
    for i in range(0, len(cityCoords)):
        if key in cityCoords[i]:
            name = (cityCoords[i])[key]
            break
    return name

def return_animal_facts(name):
    if(name == "Sea otters"): 
        x = ("Location: Found mostly on western coast of Canada, along British Columbia \n")
        x += ("Diet: sea urchins, clams, mussels, crabs \n")
        x += ("Why They're Endangered:Found mostly on western coast of Canada, along British Columbia \n")
        return x, "animals/sea_otter.jpg"
    if(name == "Blue whale"):
        x = ("Location: Found on western and eastern coast of Canada. Globally they live in all the oceans except for Atlantic \n")
        x += ("Diet: Krill, plankton and other tiny crustaceans\n")
        x += ("Why They're Endangered: Due to habitat loss and toxic materials, commercial fishing gear\n")
        return x, "animals/blue_whale.jpg"
    if(name == "Basking shark"):
        x = ("Location: Mostly found in the white bay and Notre Dame Bay, to the Gulf of St. Laurent\n")
        x += ("Diet: Plankton and other small crustaceans \n")
        x += ("Why They're Endangered: Due to collisions with boats, commercial fishing lines, and an eradication program that last for a few years in the 1950’s\n")
        return x, "animals/basking_shark.jpg"
    if(name == "Humpback whale"):
        x = ("Location: Mostly found in east and west coasts. Extending to labrador in the east\nand Northwestern Alaska in the west")
        x += ("\n\nDiet: small fish, krill, and other small crustaceans")
        x += ("\n\nWhy They're Endangered: Due to commercial whaling since the 1970’s\n\n\n")
        return x, "animals/humbackwhale.png"
    if(name == "Leatherback turtle"):
        x = ("Location: Found mostly in the Atlantic ocean on the east coast of Canada, especially along the coast of Newfoundland & Labrador \n")
        x += ("Diet: Pelagic, jellyfish, and tunicates\n")
        x += ("Why They're Endangered: Due to Commercial fishing, illegal collection of eggs, pollution, and climate change\n")
        return x, "animals/leatherback_turtle.jpg"
    if(name == "Shortnose sturgeon"):
        x = ("Location: Rivers and lakes on the east side of Canada such as the Saint John\n")
        x += ("Diet: insects, crustaceans, worms, mollusks\n")
        x += ("Why They're Endangered: Due to Habitat degradation, water pollution, dredging, commercial fishing\n")
        return x, "animals/shortnose_sturgeon.png"
    if(name == "Atlantic salmon"):
        x = ("Location: Atlantic salmon are found in various rivers and coastal areas \nalong the Atlantic coast of Canada, including the Atlantic provinces and \nparts of Quebec.")
        x += ("\n\nDiet: Their diet primarily consists of smaller fish, invertebrates, \nand insects.\n\n\n")
        x += ("\n\nWhy They're Endangered: Atlantic salmon populations are endangered due \nto habitat destruction, including dam construction, water pollution, overfishing, \nand the impact of aquaculture practices. These factors have contributed to \nthe decline of wild Atlantic salmon stocks.\n\n\n")
        return x, "animals/atlantic_salmon.png"
    if(name == "Lake sturgeon"):
        x = ("Location: Lake sturgeon are native to various watersheds across Canada, \nincluding the Great Lakes, St. Lawrence River, and many other rivers and lakes.")
        x += ("\n\nDiet: They are bottom-feeders and primarily consume aquatic \ninvertebrates, small fish, and plants.")
        x += ("\n\nLake sturgeon are endangered due to habitat loss and degradation \nfrom factors such as dam construction, pollution, and overharvesting. Their slow growth and late \nmaturity make them particularly vulnerable to population declines.\n\n\n")
        return x, "animals/lake_sturgeon.jpg"
    if(name == "White sturgeon"):
        x = ("Location: White sturgeon are found in the Fraser River and its \ntributaries in British Columbia.")
        x += ("\n\nDiet: They primarily feed on aquatic invertebrates and small fish.")
        x += ("\n\nWhy They're Endangered: White sturgeon face threats from habitat degradation, including dams\n and habitat alteration, as well as pollution and overharvesting. Their status is also \nimpacted by their slow growth and late maturity.\n\n\n")
        return x, "animals/white_sturgeo.jpg"
    if(name == "Eastern Sand darter"):
        x = ("Location: Eastern sand darters inhabit streams and rivers in the Great Lakes\n and St. Lawrence River regions.")
        x += ("\n\nDiet: They primarily feed on small invertebrates and insect larvae.")
        x += ("\n\nWhy They're Endangered: These darters are endangered due to habitat\n loss and water quality issues caused by urbanization, agriculture, and other human activities. \nFragmentation of their habitat further threatens their populations.\n\n\n")
        return x, "animals/eastern_sand_darter.jpg"
    if(name == "Vancouver lamprey"):
        x = ("Location: Vancouver lampreys are found in freshwater systems on Vancouver\n Island and the lower Fraser River in British Columbia.")
        x += ("\n\nDiet: Lampreys are parasitic and feed on the bodily fluids of other fish.")
        x += ("\n\nWhy They're Endangered: Habitat loss and water quality issues in their\n limited range have led to the decline of Vancouver lamprey populations. These \nlampreys are also sensitive to changes in water temperature and quality.\n\n\n")
        return x, "animals/vancouver_landprey.jpg"
    if(name == "Nooksack Dace"):
        x = ("Location: Nooksack dace inhabit streams and rivers in British Columbia \nand some parts of the western United States.")
        x += ("\n\nDiet: They primarily feed on aquatic invertebrates.")
        x += ("\n\nWhy They're Endangered: Habitat loss and degradation caused by urban \ndevelopment, agriculture, and water diversion have led to the decline of Nooksack \ndace populations. These factors have limited their suitable habitat.\n\n\n")
        return x, "animals/nooksack_dace.png"